from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional

import requests
from openpyxl import load_workbook
from pydantic import BaseModel

UPSTREAM_URL = "http://10.13.55.228:5004/api/outPut/exportMaterialStockToExcel"


class SWHInventoryItem(BaseModel):
    PKG_ID: Optional[str] = None
    PN: Optional[str] = None
    QTY: Optional[int] = None
    POSITION_CODE: Optional[str] = None
    AREA_CODE: Optional[str] = None


def get_swh_inventory(
    *,
    upstream_url: str = UPSTREAM_URL,
    timeout_seconds: int = 60,
    body: Optional[Dict[str, Any]] = None,
) -> List[SWHInventoryItem]:
    payload = body or {
        "reelId": "",
        "pn": "",
        "lot": "",
        "positionCode": "",
        "areaCode": "",
        "boxCode": "",
        "tray": "",
        "status": [],
        "createStartTime": None,
        "createEndTime": None,
    }

    res = requests.post(upstream_url, json=payload, timeout=timeout_seconds)
    if res.status_code < 200 or res.status_code >= 300:
        raise RuntimeError(f"Upstream error: {res.status_code}")

    workbook = load_workbook(filename=BytesIO(res.content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = rows[1:]

    mapping = {
        "ReelId": "PKG_ID",
        "PN": "PN",
        "Qty": "QTY",
        "PositionCode": "POSITION_CODE",
        "AreaCode": "AREA_CODE",
    }

    items: List[SWHInventoryItem] = []
    for row in data_rows:
        record: Dict[str, Any] = {}
        for excel_key, output_key in mapping.items():
            try:
                idx = headers.index(excel_key)
            except ValueError:
                record[output_key] = None
                continue
            value = row[idx] if idx < len(row) else None
            record[output_key] = value
        items.append(SWHInventoryItem(**record))

    return items
